# -*- coding: utf-8 -*-
import base64
import openpyxl
import io
import csv
from datetime import datetime
import logging
from odoo import models, fields, _, api
from odoo.exceptions import ValidationError, UserError

_logger = logging.getLogger(__name__)

class BankStatementImport(models.TransientModel):
    _name = 'om.bank.statement.import'
    _description = 'Import Bank Statement'

    journal_id = fields.Many2one('account.journal', string='Journal', required=True, domain=[('type', '=', 'bank')])
    file_data = fields.Binary(string='File', required=True)
    file_name = fields.Char(string='File Name')
    mapping_line_ids = fields.One2many('om.bank.statement.import.mapping', 'wizard_id', string='Column Mapping')
    
    # File options
    sheet_name = fields.Char(string='Sheet Name')
    sheet_options = fields.Selection(selection='_get_sheet_options', string='Select Sheet')
    has_header = fields.Boolean(string='Use first row as header', default=True)

    # Formatting options
    encoding = fields.Selection([
        ('utf-8', 'UTF-8'),
        ('utf-16', 'UTF-16'),
        ('windows-1252', 'Windows-1252'),
        ('latin1', 'Latin1'),
    ], string='Encoding', default='utf-8')
    
    separator = fields.Selection([
        ('comma', 'Comma'),
        ('semicolon', 'Semicolon'),
        ('tab', 'Tab'),
        ('space', 'Space'),
    ], string='Separator', default='comma')
    
    quote_char = fields.Char(string='Text Delimiter', default='"')
    
    date_format = fields.Selection([
        ('eu_slash', 'DD/MM/YYYY (31/01/2026)'),
        ('iso_dash', 'YYYY-MM-DD (2026-01-31)'),
        ('us_slash', 'MM/DD/YYYY (01/31/2026)'),
        ('eu_dash', 'DD-MM-YYYY (31-01-2026)'),
        ('eu_dot', 'DD.MM.YYYY (31.01.2026)'),
        ('iso_slash', 'YYYY/MM/DD (2026/01/31)'),
        ('eu_short', 'DD/MM/YY (31/01/26)'),
        ('us_short', 'MM/DD/YY (01/31/26)'),
    ], string='Date Format', default='eu_slash', help="Select the date format used in your file")
    
    float_decimal_separator = fields.Selection([
        ('dot', 'Dot (.)'),
        ('comma', 'Comma (,)'),
    ], string='Decimal Separator', default='dot')
    
    float_thousand_separator = fields.Selection([
        ('comma', 'Comma (,)'),
        ('dot', 'Dot (.)'),
        ('space', 'Space'),
    ], string='Thousands Separator', default='comma')

    on_error = fields.Selection([
        ('fail', 'Stop Import'),
        ('skip', 'Skip Row'),
    ], string='On Error', default='fail')
    
    create_partner = fields.Boolean(string='Create New Partners', default=True, help="If checked, new partners will be created if not found by name.")

    @api.onchange('file_data', 'sheet_options', 'has_header', 'encoding', 'separator', 'quote_char')
    def _onchange_parse_file(self):
        self.mapping_line_ids = [(5, 0, 0)]  # Clear existing lines
        if not self.file_data:
            return
        
        file_name = (self.file_name or '').lower()
        
        try:
            file_data = base64.b64decode(self.file_data)
            header = []
            sample = []
            
            if file_name.endswith('.csv'):
                try:
                    data_file = io.StringIO(file_data.decode(self.encoding or 'utf-8'))
                    
                    csv_separator = ','
                    if self.separator == 'semicolon':
                        csv_separator = ';'
                    elif self.separator == 'tab':
                        csv_separator = '\t'
                    elif self.separator == 'space':
                        csv_separator = ' '
                        
                    csv_quote = self.quote_char or '"'
                    reader = csv.reader(data_file, delimiter=csv_separator, quotechar=csv_quote)
                    
                    rows = []
                    try:
                        for i in range(2):
                            rows.append(next(reader))
                    except StopIteration:
                        pass
                        
                    if rows:
                        if self.has_header:
                            header = rows[0]
                            sample = rows[1] if len(rows) > 1 else []
                        else:
                            header = [f"Column {i+1}" for i in range(len(rows[0]))]
                            sample = rows[0]
                except Exception as e:
                    _logger.warning(f"CSV Parse Error: {e}")

            elif file_name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
                
                # Sheet selection logic
                sheet_names = wb.sheetnames
                if not self.sheet_options or self.sheet_options not in sheet_names:
                    # Default to first sheet if not selected or invalid
                    sheet = wb.worksheets[0]
                else:
                    sheet = wb[self.sheet_options]
                
                rows = list(sheet.iter_rows(values_only=True, max_row=2))
                
                if rows:
                    row1 = list(rows[0]) if rows else []
                    row2 = list(rows[1]) if len(rows) > 1 else []
                    
                    if self.has_header:
                        header = row1
                        sample = row2
                    else:
                        header = [f"Column {i+1}" for i in range(len(row1))]
                        sample = row1
                         
            lines = []
            for idx, col_name in enumerate(header):
                if not col_name and self.has_header:
                    continue
                
                col_label = str(col_name) if col_name is not None else f"Column {idx+1}"
                
                # Auto-guess field
                target = False
                col_lower = col_label.lower()
                if 'date' in col_lower:
                    target = 'date'
                elif 'amount' in col_lower or 'debit' in col_lower or 'credit' in col_lower:
                    target = 'amount'
                elif 'partner' in col_lower or 'customer' in col_lower or 'vendor' in col_lower:
                    target = 'partner'
                elif 'label' in col_lower or 'desc' in col_lower or 'ref' in col_lower:
                    target = 'payment_ref'
                elif 'curr' in col_lower:
                    target = 'foreign_currency_code'
                    
                example = str(sample[idx]) if idx < len(sample) and sample[idx] is not None else ''
                
                lines.append((0, 0, {
                    'column_index': idx,
                    'column_name': col_label,
                    'example_content': example,
                    'target_field': target
                }))
            
            self.mapping_line_ids = lines
            
        except Exception as e:
             _logger.error(f"Error parsing file for preview: {e}")
             pass

    def _get_sheet_options(self):
        if not self.file_data:
            return []
        
        if not (self.file_name or '').lower().endswith('.xlsx'):
             return []
             
        try:
             raw_data = base64.b64decode(self.file_data)
             wb = openpyxl.load_workbook(io.BytesIO(raw_data), read_only=True)
             return [(name, name) for name in wb.sheetnames]
        except:
             return []
    
    def test_import(self):
        return self.import_file(dry_run=True)

    def import_file(self, dry_run=False):
        if not self.journal_id.suspense_account_id:
            raise ValidationError(_("The journal '%s' does not have a Suspense Account defined. Please go to Accounting/Invoicing Configuration -> Journals and set a Suspense Account for this journal.") % self.journal_id.name)

        if not self.mapping_line_ids:
             raise ValidationError(_("Please map columns before importing. Re-upload file to refresh mapping."))
             
        # Validate unique mappings
        mapped_targets = [line.target_field for line in self.mapping_line_ids if line.target_field]
        if len(mapped_targets) != len(set(mapped_targets)):
             import collections
             duplicates = [item for item, count in collections.Counter(mapped_targets).items() if count > 1]
             raise ValidationError(_("Duplicate mapping detected for fields: %s. Please map each Odoo Field to only one column.") % ", ".join(duplicates))

        # Validate critical mappings
        mapping = {line.target_field: line.column_index for line in self.mapping_line_ids if line.target_field}
        if 'date' not in mapping:
             raise ValidationError(_("Please map a 'Date' column."))
        if 'amount' not in mapping:
             raise ValidationError(_("Please map an 'Amount' column."))

        statement = False
        file_name = (self.file_name or '').lower()
        raw_data = base64.b64decode(self.file_data)
        
        logs = []
        
        if file_name.endswith('.csv'):
            res = self._import_csv(raw_data, mapping, dry_run=dry_run)
            if dry_run:
                logs = res
            else:
                statement = res
        elif file_name.endswith('.xlsx'):
            res = self._import_xlsx(raw_data, mapping, dry_run=dry_run)
            if dry_run:
                logs = res
            else:
                statement = res
        else:
            raise ValidationError(_("Invalid file format. Please upload .csv or .xlsx file."))
        
        if dry_run:
             message = "\n".join(logs)
             is_success = True
             if not logs:
                 title = _("Test Failed")
                 message = "No logs generated."
                 is_success = False
             elif "Error" not in message and "Fail" not in message:
                 title = _("Test Successful!")
                 is_success = True
             else:
                 title = _("Test Completed with Issues")
                 is_success = False
                 
             return {
                 'type': 'ir.actions.client',
                 'tag': 'display_notification',
                 'params': {
                     'title': title,
                     'message': message,
                     'type': 'success' if is_success else 'warning',
                     'sticky': True,
                 }
             }

        if statement:
            return {
                'type': 'ir.actions.act_window',
                'name': _('Bank Statement'),
                'view_mode': 'form',
                'res_model': 'account.bank.statement',
                'res_id': statement.id,
                'target': 'current',
            }

    def _parse_float(self, value):
        if not value:
            return 0.0
        if isinstance(value, float) or isinstance(value, int):
            return float(value)
        
        value = str(value).strip()
        
        thousand_sep = ','
        if self.float_thousand_separator == 'dot':
            thousand_sep = '.'
        elif self.float_thousand_separator == 'space':
            thousand_sep = ' '
            
        decimal_sep = '.'
        if self.float_decimal_separator == 'comma':
            decimal_sep = ','
            
        if thousand_sep:
            value = value.replace(thousand_sep, '')
        if decimal_sep != '.':
            value = value.replace(decimal_sep, '.')
            
        try:
            return float(value)
        except ValueError:
             # Let caller handle error
             raise ValueError("Invalid Float")

    def _parse_date(self, value):
        if not value:
             return False
        
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, (int, float)): # Excel serial date
             pass

        # Map selection key to actual format string
        format_map = {
            'iso_dash': '%Y-%m-%d',
            'eu_slash': '%d/%m/%Y',
            'us_slash': '%m/%d/%Y',
            'eu_dash': '%d-%m-%Y',
            'eu_dot': '%d.%m.%Y',
            'iso_slash': '%Y/%m/%d',
            'eu_short': '%d/%m/%y',
            'us_short': '%m/%d/%y',
        }
        date_fmt = format_map.get(self.date_format, '%Y-%m-%d')

        value = str(value).strip()
        try:
            return datetime.strptime(value, date_fmt).date()
        except ValueError:
            # Fallback to standard ISO
             try:
                 return datetime.strptime(value, '%Y-%m-%d').date()
             except:
                 raise ValueError("Invalid Date")

    def _import_csv(self, raw_data, mapping, dry_run=False):
        logs = []
        try:
            data_file = io.StringIO(raw_data.decode(self.encoding or 'utf-8'))
            
            csv_separator = ','
            if self.separator == 'semicolon':
                csv_separator = ';'
            elif self.separator == 'tab':
                csv_separator = '\t'
            elif self.separator == 'space':
                csv_separator = ' '
            
            csv_quote = self.quote_char or '"'
            
            reader = csv.reader(data_file, delimiter=csv_separator, quotechar=csv_quote)
            
            statement_vals = {
                'name': self.file_name or 'Imported Statement',
                'journal_id': self.journal_id.id,
                'line_ids': [],
            }
            
            if self.has_header:
                 try:
                    next(reader) 
                 except StopIteration:
                     pass

            row_idx = 1 if self.has_header else 0
            valid_rows = 0
            skipped_rows = 0
            
            for row in reader:
                row_idx += 1
                if not row:
                    continue
                
                try:
                    vals = self._extract_values(row, mapping)
                    if not dry_run:
                        statement_vals['line_ids'].append((0, 0, vals))
                    valid_rows += 1
                except Exception as e:
                    skipped_rows += 1
                    msg = f"Row {row_idx}: {str(e)}"
                    logs.append(msg)
                    if self.on_error == 'fail':
                        raise ValidationError(msg)
            
            if dry_run:
                 summary = f"Processed {row_idx} lines.\nValid: {valid_rows}\nSkipped: {skipped_rows}"
                 logs.insert(0, summary)
                 return logs

            if statement_vals['line_ids']:
                return self.env['account.bank.statement'].create(statement_vals)
            else:
                if skipped_rows > 0:
                     raise ValidationError(_("No valid transactions found. %d lines were skipped due to errors.") % skipped_rows)
                else:
                     raise ValidationError(_("No valid transactions found."))

        except Exception as e:
            if dry_run:
                return [f"Fatal CSV Error: {str(e)}"]
            raise ValidationError(_("Error parsing CSV file: %s") % str(e))

    def _import_xlsx(self, raw_data, mapping, dry_run=False):
        logs = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_data), data_only=True)
            
            # Use selected sheet
            if self.sheet_options and self.sheet_options in wb.sheetnames:
                sheet = wb[self.sheet_options]
            else:
                sheet = wb.worksheets[0]
            
            statement_vals = {
                'name': self.file_name or 'Imported Statement',
                'journal_id': self.journal_id.id,
                'line_ids': [],
            }
            
            rows = list(sheet.iter_rows(values_only=True))
            
            start_row = 1 if self.has_header else 0
            row_idx = start_row
            valid_rows = 0
            skipped_rows = 0
            
            for row in rows[start_row:]:
                row_idx += 1
                if not row or not any(row):
                    continue
                
                try:
                    vals = self._extract_values(row, mapping)
                    if not dry_run:
                        statement_vals['line_ids'].append((0, 0, vals))
                    valid_rows += 1
                except Exception as e:
                    skipped_rows += 1
                    msg = f"Row {row_idx}: {str(e)}"
                    logs.append(msg)
                    if self.on_error == 'fail':
                         raise ValidationError(msg)

            if dry_run:
                 summary = f"Processed {len(rows)} lines.\nValid: {valid_rows}\nSkipped: {skipped_rows}"
                 logs.insert(0, summary)
                 return logs

            if statement_vals['line_ids']:
                return self.env['account.bank.statement'].create(statement_vals)
            else:
                 if skipped_rows > 0:
                     raise ValidationError(_("No valid transactions found. %d lines were skipped due to errors.") % skipped_rows)
                 else:
                     raise ValidationError(_("No valid transactions found."))

        except Exception as e:
            if dry_run:
                return [f"Fatal XLSX Error: {str(e)}"]
            raise ValidationError(_("Error parsing XLSX file: %s") % str(e))

    def _extract_values(self, row, mapping):
        # row is a list/tuple of values
        
        # 1. Date
        date_idx = mapping.get('date')
        date_val = row[date_idx] if date_idx < len(row) else False
        try:
             date_obj = self._parse_date(date_val)
             if not date_obj:
                 date_obj = fields.Date.today()
        except ValueError as e:
             # Add context about column
             col_name = "Unknown"
             for line in self.mapping_line_ids:
                 if line.column_index == date_idx:
                     col_name = line.column_name
                     break
             raise UserError(f"Date Error: '{date_val}' in column '{col_name}' (Index {date_idx}) - {str(e)}")

        # 2. Label
        label_idx = mapping.get('payment_ref')
        label_val = row[label_idx] if label_idx is not None and label_idx < len(row) else '/'
        label_val = str(label_val) if label_val else '/'
        
        # 3. Partner
        partner_id = False
        partner_idx = mapping.get('partner')
        if partner_idx is not None and partner_idx < len(row):
             partner_name = str(row[partner_idx]).strip() if row[partner_idx] else ''
             partner_id = self._find_or_create_partner(partner_name)

        # 4. Amount
        amount_idx = mapping.get('amount')
        amount_val = row[amount_idx] if amount_idx < len(row) else 0.0
        try:
             amount = self._parse_float(amount_val)
        except ValueError:
             col_name = "Unknown"
             for line in self.mapping_line_ids:
                 if line.column_index == amount_idx:
                     col_name = line.column_name
                     break
             raise UserError(f"Amount Error: '{amount_val}' in column '{col_name}' (Index {amount_idx})")
        
        # 5. Currency
        foreign_currency_id = False
        amount_currency = 0.0
        
        currency_code_idx = mapping.get('foreign_currency_code')
        amount_currency_idx = mapping.get('amount_currency')
        
        if currency_code_idx is not None and currency_code_idx < len(row):
            code = str(row[currency_code_idx]).strip()
            if code:
                currency = self.env['res.currency'].search([('name', '=', code)], limit=1)
                if currency:
                    foreign_currency_id = currency.id
                
        if foreign_currency_id:
             if amount_currency_idx is not None and amount_currency_idx < len(row):
                  try:
                       amount_currency = self._parse_float(row[amount_currency_idx])
                  except ValueError:
                       amount_currency = 0.0
             else:
                  # If no specific foreign amount column, assume main amount is foreign
                  pass 

        vals = {
            'date': date_obj,
            'payment_ref': label_val,
            'partner_id': partner_id,
            'amount': amount,
            'journal_id': self.journal_id.id,
        }
        
        if foreign_currency_id:
             vals['foreign_currency_id'] = foreign_currency_id
             vals['amount_currency'] = amount_currency
             
        return vals

    def _find_or_create_partner(self, name):
        if not name:
            return False
        # Case insensitive exact match or close match
        partner = self.env['res.partner'].search([('name', '=ilike', name)], limit=1)
        if partner:
            return partner.id
        elif self.create_partner:
             # Create new partner
             try:
                 new_partner = self.env['res.partner'].create({'name': name, 'type': 'contact'})
                 return new_partner.id
             except Exception as e:
                 _logger.warning(f"Failed to create partner {name}: {e}")
                 return False
        else:
             return False

class BankStatementImportMapping(models.TransientModel):
    _name = 'om.bank.statement.import.mapping'
    _description = 'Bank Import Mapping'

    wizard_id = fields.Many2one('om.bank.statement.import', string='Wizard')
    column_index = fields.Integer(string='Column Index')
    column_name = fields.Char(string='Column Name')
    example_content = fields.Char(string='Example Content')
    target_field = fields.Selection([
        ('date', 'Date'),
        ('payment_ref', 'Label'),
        ('partner', 'Partner'),
        ('amount', 'Amount'),
        ('foreign_currency_code', 'Foreign Currency Code'),
        ('amount_currency', 'Foreign Currency Amount'),
    ], string='Odoo Field')


